from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.image_service import (
    build_image_signature,
    clean_row_for_search,
    compare_signature,
    extract_images_by_row,
)
from utils.db import BASE_PATH, ALLOWED_EXTENSIONS, get_db_connection
from utils.helpers import (
    STATE,
    STANDARD_HEADERS,
    detect_header_row,
    get_public_imports_state,
    get_public_workbook_state,
    infer_partition_from_sheets,
    json_loads_safe,
    merge_extra_data_into_row,
    normalize_cell_value,
    normalize_excel_header,
    rebuild_global_search_index,
    row_has_meaningful_data,
)


def _rows_to_dicts(cursor, rows):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in rows]


def _fetchall_dict(cursor):
    rows = cursor.fetchall()
    if not rows:
        return []
    return _rows_to_dicts(cursor, rows)


def _fetchone_dict(cursor):
    row = cursor.fetchone()
    if not row:
        return None
    columns = [col[0] for col in cursor.description]
    return dict(zip(columns, row))


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS



def parse_excel_file(file_path: Path) -> dict[str, Any]:
    import uuid

    import_id = uuid.uuid4().hex
    workbook = load_workbook(file_path, data_only=True)
    all_sheets: list[dict[str, Any]] = []
    total_rows = 0
    search_index: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header_row = detect_header_row(ws)
        headers: list[str] = []
        column_mappings: list[tuple[str, int]] = []
        seen_headers: set[str] = set()

        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(header_row, col_idx).value
            if header_value in (None, ""):
                continue

            normalized_header = normalize_excel_header(str(header_value).strip())
            column_mappings.append((normalized_header, col_idx))
            if normalized_header not in seen_headers:
                seen_headers.add(normalized_header)
                headers.append(normalized_header)

        if not headers:
            continue

        images_by_row = extract_images_by_row(ws, import_id=import_id, sheet_name=sheet_name)
        rows: list[dict[str, Any]] = []

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data: dict[str, Any] = {}
            for header, col_idx in column_mappings:
                cell_value = normalize_cell_value(ws.cell(row_idx, col_idx).value)
                if header not in row_data:
                    row_data[header] = cell_value
                else:
                    old_value = row_data.get(header)
                    if old_value in (None, "", []) and cell_value not in (None, "", []):
                        row_data[header] = cell_value

            row_images = images_by_row.get(row_idx, [])
            row_data["__images"] = [item["src"] for item in row_images]
            row_data["__excelRow"] = row_idx
            row_data["__folderId"] = None
            row_data["__folderName"] = None
            row_data["__detailUrl"] = None

            has_meaningful_data = row_has_meaningful_data(row_data, headers)
            if row_images:
                has_meaningful_data = True

            if has_meaningful_data:
                rows.append(row_data)
                for image_item in row_images:
                    search_index.append(
                        {
                            "sheetName": sheet_name,
                            "excelRow": row_idx,
                            "rowRef": row_data,
                            "matchedImage": image_item["src"],
                            "hash": image_item["hash"],
                            "compare_png": image_item["compare_png"],
                        }
                    )

        total_rows += len(rows)
        all_sheets.append(
            {
                "sheetName": sheet_name,
                "headerRow": header_row,
                "headers": headers,
                "rowCount": len(rows),
                "rows": rows,
            }
        )

    partition_info = infer_partition_from_sheets(all_sheets)
    workbook_data = {
        "id": import_id,
        "fileName": file_path.name,
        "sheetCount": len(all_sheets),
        "totalRows": total_rows,
        "imageIndexCount": len(search_index),
        "mappedFolderCount": 0,
        "folderImportName": None,
        "customer": partition_info.get("customer"),
        "season": partition_info.get("season"),
        "customerValues": partition_info.get("customerValues", []),
        "seasonValues": partition_info.get("seasonValues", []),
        "isAmbiguous": partition_info.get("isAmbiguous", False),
        "sheets": all_sheets,
        "search_index": search_index,
    }

    STATE["imports"].append(workbook_data)
    rebuild_global_search_index()
    return {k: v for k, v in workbook_data.items() if k != "search_index"}



def save_workbook_to_sqlserver(workbook_data: dict[str, Any]) -> None:
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            """
            INSERT INTO excel_imports (id, file_name, sheet_count, total_rows, image_index_count)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                workbook_data["id"],
                workbook_data["fileName"],
                workbook_data["sheetCount"],
                workbook_data["totalRows"],
                workbook_data["imageIndexCount"],
            ),
        )

        for sheet in workbook_data.get("sheets", []):
            cursor.execute(
                """
                SET NOCOUNT ON;
                INSERT INTO excel_sheets (import_id, sheet_name, header_row, row_count)
                VALUES (?, ?, ?, ?);
                SELECT SCOPE_IDENTITY() AS id;
                """,
                (
                    workbook_data["id"],
                    sheet["sheetName"],
                    sheet.get("headerRow"),
                    sheet.get("rowCount", 0),
                ),
            )
            sheet_id_row = cursor.fetchone()
            sheet_id = int(sheet_id_row[0])

            for row in sheet.get("rows", []):
                extra_json = {
                    k: v for k, v in row.items()
                    if not k.startswith("__") and k not in STANDARD_HEADERS
                }

                row_no = row.get("No")
                if isinstance(row_no, str) and row_no.isdigit():
                    row_no = int(row_no)

                cursor.execute(
                    """
                    SET NOCOUNT ON;
                    INSERT INTO pattern_rows (
                        import_id, sheet_id, excel_row,
                        row_no, customer, season, staff,
                        style_no, style_name, product, categories, gender,
                        extra_json, folder_id, folder_name, detail_url
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                    SELECT SCOPE_IDENTITY() AS id;
                    """,
                    (
                        workbook_data["id"],
                        sheet_id,
                        row.get("__excelRow"),
                        row_no,
                        row.get("Customer"),
                        row.get("Season"),
                        row.get("Staff"),
                        row.get("Style No"),
                        row.get("Style Name"),
                        row.get("Product"),
                        row.get("Categories"),
                        row.get("Gender"),
                        json.dumps(extra_json, ensure_ascii=False) if extra_json else None,
                        row.get("__folderId"),
                        row.get("__folderName"),
                        row.get("__detailUrl"),
                    ),
                )
                pattern_row_id_row = cursor.fetchone()
                pattern_row_id = int(pattern_row_id_row[0])

                image_records = [
                    item
                    for item in workbook_data.get("search_index", [])
                    if item["rowRef"] is row
                ]

                for idx, image_item in enumerate(image_records):
                    cursor.execute(
                        """
                        INSERT INTO pattern_row_images (
                            pattern_row_id, image_order, image_src, image_hash, compare_png
                        )
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            pattern_row_id,
                            idx,
                            image_item["matchedImage"],
                            str(image_item["hash"]),
                            image_item["compare_png"],
                        ),
                    )

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


def restore_state_from_sqlserver() -> None:
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        STATE["imports"] = []
        STATE["search_index"] = []
        STATE["folder_index"] = {}
        STATE["folder_root_dir"] = []

        cursor.execute(
            """
            SELECT id, file_name, sheet_count, total_rows, image_index_count
            FROM excel_imports
            ORDER BY id ASC
            """
        )
        import_rows = _fetchall_dict(cursor)
        if not import_rows:
            return

        cursor.execute(
            """
            SELECT id, import_id, sheet_name, header_row, row_count
            FROM excel_sheets
            ORDER BY import_id ASC, id ASC
            """
        )
        sheet_rows = _fetchall_dict(cursor)

        sheets_by_import: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for sheet in sheet_rows:
            sheet_item = {
                "id": sheet["id"],
                "import_id": sheet["import_id"],
                "sheetName": sheet["sheet_name"],
                "headerRow": sheet["header_row"],
                "rowCount": sheet["row_count"] or 0,
                "headers": [],
                "rows": [],
            }
            sheets_by_import[sheet["import_id"]].append(sheet_item)

        sheet_obj_by_id: dict[int, dict[str, Any]] = {}
        for sheets in sheets_by_import.values():
            for sheet in sheets:
                sheet_obj_by_id[sheet["id"]] = sheet

        cursor.execute(
            """
            SELECT
                id, import_id, sheet_id, excel_row, row_no, customer, season, staff,
                style_no, style_name, product, categories, gender,
                extra_json, folder_id, folder_name, detail_url
            FROM pattern_rows
            ORDER BY import_id ASC, sheet_id ASC, excel_row ASC, id ASC
            """
        )
        row_rows = _fetchall_dict(cursor)
        row_obj_by_id: dict[int, dict[str, Any]] = {}

        for db_row in row_rows:
            row_data: dict[str, Any] = {
                "No": db_row.get("row_no"),
                "Customer": db_row.get("customer"),
                "Season": db_row.get("season"),
                "Staff": db_row.get("staff"),
                "Style No": db_row.get("style_no"),
                "Style Name": db_row.get("style_name"),
                "Product": db_row.get("product"),
                "Categories": db_row.get("categories"),
                "Gender": db_row.get("gender"),
                "__images": [],
                "__excelRow": db_row.get("excel_row"),
                "__folderId": db_row.get("folder_id"),
                "__folderName": db_row.get("folder_name"),
                "__detailUrl": db_row.get("detail_url"),
            }

            detail_url = row_data.get("__detailUrl")
            if detail_url and str(detail_url).startswith("/folder/"):
                row_data["__detailUrl"] = f"{BASE_PATH}{detail_url}"

            merge_extra_data_into_row(
                row_data,
                json_loads_safe(db_row.get("extra_json")),
            )

            sheet_obj = sheet_obj_by_id.get(db_row["sheet_id"])
            if not sheet_obj:
                continue

            sheet_obj["rows"].append(row_data)
            row_obj_by_id[db_row["id"]] = row_data

        cursor.execute(
            """
            SELECT pattern_row_id, image_order, image_src, image_hash, compare_png
            FROM pattern_row_images
            ORDER BY pattern_row_id ASC, image_order ASC, id ASC
            """
        )
        image_rows = _fetchall_dict(cursor)
        rebuilt_search_index: list[dict[str, Any]] = []

        row_id_to_sheet_name: dict[int, str] = {}
        cursor.execute(
            """
            SELECT pr.id AS pattern_row_id, es.sheet_name
            FROM pattern_rows pr
            INNER JOIN excel_sheets es ON pr.sheet_id = es.id
            ORDER BY pr.id ASC
            """
        )
        join_rows = _fetchall_dict(cursor)
        for item in join_rows:
            row_id_to_sheet_name[item["pattern_row_id"]] = item["sheet_name"]

        for image_row in image_rows:
            row_obj = row_obj_by_id.get(image_row["pattern_row_id"])
            if not row_obj:
                continue

            image_src = image_row.get("image_src")
            if image_src and str(image_src).startswith("/media/"):
                image_src = f"{BASE_PATH}{image_src}"

            if image_src:
                row_obj["__images"].append(image_src)

            compare_png = image_row.get("compare_png")
            if compare_png is not None and not isinstance(compare_png, (bytes, bytearray)):
                try:
                    compare_png = bytes(compare_png)
                except Exception:
                    compare_png = None

            rebuilt_search_index.append(
                {
                    "sheetName": row_id_to_sheet_name.get(image_row["pattern_row_id"], ""),
                    "excelRow": row_obj.get("__excelRow"),
                    "rowRef": row_obj,
                    "matchedImage": image_src,
                    "hash": int(image_row["image_hash"]) if image_row.get("image_hash") not in (None, "") else 0,
                    "compare_png": compare_png,
                }
            )

        known_header_order = [
            "No",
            "Customer",
            "Season",
            "Staff",
            "Style No",
            "Style Name",
            "Product",
            "Categories",
            "Gender",
            "Sketch Design",
        ]

        for sheets in sheets_by_import.values():
            for sheet in sheets:
                header_set = []
                seen = set()

                for header in known_header_order:
                    real_header = "Sketch Design" if header == "Sketch Design" else header
                    if real_header == "Sketch Design":
                        has_value = any(row.get("__images") for row in sheet["rows"])
                    else:
                        has_value = any(
                            row.get(real_header) not in (None, "", [])
                            for row in sheet["rows"]
                        )

                    if has_value and real_header not in seen:
                        seen.add(real_header)
                        header_set.append(real_header)

                for row in sheet["rows"]:
                    for key in row.keys():
                        if key.startswith("__"):
                            continue
                        if key not in seen:
                            seen.add(key)
                            header_set.append(key)

                if any(row.get("__images") for row in sheet["rows"]) and "Sketch Design" not in seen:
                    header_set.append("Sketch Design")

                sheet["headers"] = header_set
                sheet["rowCount"] = len(sheet["rows"])

        for import_item in import_rows:
            all_sheets = sheets_by_import.get(import_item["id"], [])
            partition_info = infer_partition_from_sheets(all_sheets)

            workbook_search_index = [
                item
                for item in rebuilt_search_index
                if any(item["rowRef"] is row for sheet in all_sheets for row in sheet["rows"])
            ]

            workbook_data = {
                "id": import_item["id"],
                "fileName": import_item["file_name"],
                "sheetCount": import_item["sheet_count"] or len(all_sheets),
                "totalRows": import_item["total_rows"] or sum(len(sheet["rows"]) for sheet in all_sheets),
                "imageIndexCount": import_item["image_index_count"] or len(workbook_search_index),
                "mappedFolderCount": sum(
                    1 for sheet in all_sheets for row in sheet["rows"] if row.get("__folderId")
                ),
                "folderImportName": None,
                "customer": partition_info.get("customer"),
                "season": partition_info.get("season"),
                "customerValues": partition_info.get("customerValues", []),
                "seasonValues": partition_info.get("seasonValues", []),
                "isAmbiguous": partition_info.get("isAmbiguous", False),
                "sheets": all_sheets,
                "search_index": workbook_search_index,
            }

            STATE["imports"].append(workbook_data)

        rebuild_global_search_index()
    finally:
        cursor.close()
        conn.close()

def search_by_uploaded_image(pasted_image) -> dict[str, Any]:
    if not STATE["search_index"]:
        return {
            "success": False,
            "message": "Bạn cần import Excel trước khi dán ảnh để tìm kiếm",
            "status": 400,
        }
    if not pasted_image:
        return {"success": False, "message": "Không nhận được ảnh từ clipboard", "status": 400}

    image_bytes = pasted_image.read()
    query_signature = build_image_signature(image_bytes)
    results = []
    for item in STATE["search_index"]:
        score = compare_signature(query_signature, item)
        row_ref = item["rowRef"]
        results.append(
            {
                "sheetName": item["sheetName"],
                "excelRow": item["excelRow"],
                "score": round(score, 4),
                "scorePercent": int(round(score * 100)),
                "matchedImage": item["matchedImage"],
                "detailUrl": row_ref.get("__detailUrl"),
                "folderName": row_ref.get("__folderName"),
                "row": clean_row_for_search(row_ref),
            }
        )
    results.sort(key=lambda x: x["score"], reverse=True)
    return {
        "success": True,
        "message": "So khớp ảnh thành công",
        "data": {"results": results[:10]},
        "status": 200,
    }
