from __future__ import annotations

import base64
import re
import shutil
import uuid
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
import zipfile

import json
import mysql.connector

from flask import Flask, abort, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from PIL import Image, ImageChops, ImageOps, ImageStat
from werkzeug.utils import secure_filename

from functools import wraps
from flask import session, redirect, url_for
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

EXCEL_IMAGE_DIR = UPLOAD_DIR / "excel_images"
EXCEL_IMAGE_DIR.mkdir(parents=True, exist_ok=True)

FOLDER_IMPORT_DIR = UPLOAD_DIR / "folder_imports"
FOLDER_IMPORT_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
PDF_EXTENSIONS = {".pdf"}

DB_CONFIG = {
    "host": "127.0.0.1",
    "user": "root",
    "password": "dinh1806",
    "database": "manual_db",
    "charset": "utf8mb4",
}

def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)

def get_current_user():
    user_id = session.get("user_id")

    if not user_id:
        return None
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary = True)

    try:
        cursor.execute(
            "SELECT * from users WHERE id = %s and is_active = 1",
            (user_id,)
        )
        user = cursor.fetchone()
        return user

    finally:
        cursor.close()
        conn.close()

def user_to_public_payload(user: dict) -> dict:
    return {
        "id": user["id"],
        "email": user["email"],
        "fullName": user.get("full_name") or "",
        "isAdmin": bool(user.get("is_admin")),
        "permissions": {
            "canImportExcel": bool(user.get("can_import_excel")),
            "canImportFolder": bool(user.get("can_import_folder")),
            "canSearchImage": bool(user.get("can_search_image")),
            "canViewData": bool(user.get("can_view_data")),
            "canResetData": bool(user.get("can_reset_data")),
            "canManageUsers": bool(user.get("can_manage_users")),
            "canViewAuditLogs": bool(user.get("can_view_audit_logs")),
        },
    }

def login_required_page(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("login_page"))
        return fn(*args, **kwargs)
    return wrapper


def login_required_api(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"success": False, "message": "Bạn chưa đăng nhập"}), 401
        return fn(*args, **kwargs)
    return wrapper


def permission_required_api(permission_name: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                return jsonify({"success": False, "message": "Bạn chưa đăng nhập"}), 401

            if not bool(user.get(permission_name)):
                return jsonify(
                    {"success": False, "message": "Bạn không có quyền thực hiện thao tác này"}
                ), 403

            return fn(*args, **kwargs)
        return wrapper
    return decorator


def permission_required_page(permission_name: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("login_page"))

            if not bool(user.get(permission_name)):
                return redirect(url_for("index"))

            return fn(*args, **kwargs)
        return wrapper
    return decorator


app = Flask(__name__)
app.secret_key = "pattern_manual_secret_key"
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500 MB

STATE: dict[str, Any] = {
    "imports": [],
    "search_index": [],
    "folder_index": {},
    "folder_root_dir": [],
}

def normalize_email(email: str) -> str:
    return str(email or "").strip().lower();

def infer_partition_from_sheets(all_sheets: list[dict[str, Any]]) -> dict[str, Any]:
    customer_values: dict[str, int] = {}
    season_values: dict[str, int] = {}

    def add_count(counter: dict[str, int], value: Any) -> None:
        text = str(value or "").strip()
        if not text:
            return
        counter[text] = counter.get(text, 0) + 1

    for sheet in all_sheets:
        for row in sheet.get("rows", []):
            add_count(customer_values, row.get("Customer"))
            add_count(season_values, row.get("Season"))

    customers = sorted(customer_values.keys())
    seasons = sorted(season_values.keys())

    dominant_customer = max(customer_values, key=customer_values.get) if customer_values else None
    dominant_season = max(season_values, key=season_values.get) if season_values else None

    return {
        "customer": dominant_customer,
        "season": dominant_season,
        "customerValues": customers,
        "seasonValues": seasons,
        "isAmbiguous": len(customers) > 1 or len(seasons) > 1,
    }


def slugify_filename(value: str) -> str:
    value = str(value or "").strip()
    value = re.sub(r"[^\w\-\.]+", "_", value, flags=re.UNICODE)
    return value.strip("_") or "sheet"

def save_excel_image_file(
    image_bytes: bytes,
    import_id: str,
    sheet_name: str,
    row_idx: int,
    image_index: int,
) -> dict[str, Any]:
    normalized = load_normalized_image_from_bytes(image_bytes)
    sheet_safe = slugify_filename(sheet_name)

    target_dir = EXCEL_IMAGE_DIR / import_id / sheet_safe / str(row_idx)
    target_dir.mkdir(parents=True, exist_ok=True)

    file_name = f"image_{image_index + 1}.png"
    file_path = target_dir / file_name

    normalized.save(file_path, format="PNG", optimize=True)

    relative_path = file_path.relative_to(UPLOAD_DIR).as_posix()

    return {
        "src": f"/media/{relative_path}",
        "relative_path": relative_path,
        "hash": compute_dhash(normalized),
        "compare_png": build_compare_gray_png(normalized),
    }

def save_workbook_to_mysql(workbook_data: dict[str, Any]) -> None:
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            """
            INSERT INTO excel_imports (id, file_name, sheet_count, total_rows, image_index_count)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                workbook_data["id"],
                workbook_data["fileName"],
                workbook_data["sheetCount"],
                workbook_data["totalRows"],
                workbook_data["imageIndexCount"],
            ),
        )

        for sheet in workbook_data.get("sheets", []):
            cursor.execute(
                """
                INSERT INTO excel_sheets (import_id, sheet_name, header_row, row_count)
                VALUES (%s, %s, %s, %s)
                """,
                (
                    workbook_data["id"],
                    sheet["sheetName"],
                    sheet.get("headerRow"),
                    sheet.get("rowCount", 0),
                ),
            )
            sheet_id = cursor.lastrowid

            headers = set(sheet.get("headers", []))

            for row in sheet.get("rows", []):
                known_keys = STANDARD_HEADERS

                extra_json = {
                    k: v for k, v in row.items()
                    if not k.startswith("__") and k not in known_keys
                }

                row_no = row.get("No")
                if isinstance(row_no, str) and row_no.isdigit():
                    row_no = int(row_no)

                cursor.execute(
                    """
                    INSERT INTO pattern_rows (
                        import_id, sheet_id, excel_row,
                        row_no, customer, season, staff,
                        style_no, style_name, product, categories, gender,
                        extra_json, folder_id, folder_name, detail_url
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        workbook_data["id"],
                        sheet_id,
                        row.get("__excelRow"),
                        row_no,
                        row.get("Customer"),
                        row.get("Season"),
                        row.get("Staff"),
                        row.get("Style No"),
                        row.get("Style Name"),
                        row.get("Product"),
                        row.get("Categories"),
                        row.get("Gender"),
                        json.dumps(extra_json, ensure_ascii=False) if extra_json else None,
                        row.get("__folderId"),
                        row.get("__folderName"),
                        row.get("__detailUrl"),
                    ),
                )
                pattern_row_id = cursor.lastrowid

                images = row.get("__images") or []
                image_records = [
                    item for item in workbook_data.get("search_index", [])
                    if item["rowRef"] is row
                ]

                for idx, image_item in enumerate(image_records):
                    cursor.execute(
                        """
                        INSERT INTO pattern_row_images (
                            pattern_row_id, image_order, image_src, image_hash, compare_png
                        )
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (
                            pattern_row_id,
                            idx,
                            image_item["matchedImage"],
                            str(image_item["hash"]),
                            image_item["compare_png"],
                        ),
                    )

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

def restore_state_from_mysql() -> None:
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # reset state trước khi rebuild
        STATE["imports"] = []
        STATE["search_index"] = []
        STATE["folder_index"] = {}
        STATE["folder_root_dir"] = None

        # 1) lấy toàn bộ import
        cursor.execute(
            """
            SELECT
                id,
                file_name,
                sheet_count,
                total_rows,
                image_index_count
            FROM excel_imports
            ORDER BY id ASC
            """
        )
        import_rows = cursor.fetchall() or []

        if not import_rows:
            return

        # 2) lấy toàn bộ sheet
        cursor.execute(
            """
            SELECT
                id,
                import_id,
                sheet_name,
                header_row,
                row_count
            FROM excel_sheets
            ORDER BY import_id ASC, id ASC
            """
        )
        sheet_rows = cursor.fetchall() or []

        # gom sheet theo import
        sheets_by_import: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for sheet in sheet_rows:
            sheet_item = {
                "id": sheet["id"],
                "import_id": sheet["import_id"],
                "sheetName": sheet["sheet_name"],
                "headerRow": sheet["header_row"],
                "rowCount": sheet["row_count"] or 0,
                "headers": [],
                "rows": [],
            }
            sheets_by_import[sheet["import_id"]].append(sheet_item)

        # map sheet_id -> object sheet để gắn row
        sheet_obj_by_id: dict[int, dict[str, Any]] = {}
        for import_id, sheets in sheets_by_import.items():
            for sheet in sheets:
                sheet_obj_by_id[sheet["id"]] = sheet

        # 3) lấy toàn bộ rows
        cursor.execute(
            """
            SELECT
                id,
                import_id,
                sheet_id,
                excel_row,
                row_no,
                customer,
                season,
                staff,
                style_no,
                style_name,
                product,
                categories,
                gender,
                extra_json,
                folder_id,
                folder_name,
                detail_url
            FROM pattern_rows
            ORDER BY import_id ASC, sheet_id ASC, excel_row ASC, id ASC
            """
        )
        row_rows = cursor.fetchall() or []

        # map pattern_row_id -> row object
        row_obj_by_id: dict[int, dict[str, Any]] = {}

        for db_row in row_rows:
            extra_json = db_row.get("extra_json")
            extra_data = {}

            if extra_json:
                if isinstance(extra_json, str):
                    try:
                        extra_data = json.loads(extra_json)
                    except Exception:
                        extra_data = {}
                elif isinstance(extra_json, dict):
                    extra_data = extra_json

            row_data: dict[str, Any] = {
                "No": db_row.get("row_no"),
                "Customer": db_row.get("customer"),
                "Season": db_row.get("season"),
                "Staff": db_row.get("staff"),
                "Style No": db_row.get("style_no"),
                "Style Name": db_row.get("style_name"),
                "Product": db_row.get("product"),
                "Categories": db_row.get("categories"),
                "Gender": db_row.get("gender"),
                "__images": [],
                "__excelRow": db_row.get("excel_row"),
                "__folderId": db_row.get("folder_id"),
                "__folderName": db_row.get("folder_name"),
                "__detailUrl": db_row.get("detail_url"),
            }

            # gắn thêm các cột động từ extra_json
            merge_extra_data_into_row(row_data, extra_data)

            sheet_obj = sheet_obj_by_id.get(db_row["sheet_id"])
            if not sheet_obj:
                continue

            sheet_obj["rows"].append(row_data)
            row_obj_by_id[db_row["id"]] = row_data

        # 4) lấy toàn bộ image rows
        cursor.execute(
            """
            SELECT
                pattern_row_id,
                image_order,
                image_src,
                image_hash,
                compare_png
            FROM pattern_row_images
            ORDER BY pattern_row_id ASC, image_order ASC, id ASC
            """
        )
        image_rows = cursor.fetchall() or []

        # build search index từ DB
        rebuilt_search_index: list[dict[str, Any]] = []

        # cần biết row này thuộc sheet nào để trả sheetName
        row_id_to_sheet_name: dict[int, str] = {}
        cursor.execute(
            """
            SELECT
                pr.id AS pattern_row_id,
                es.sheet_name
            FROM pattern_rows pr
            INNER JOIN excel_sheets es ON pr.sheet_id = es.id
            ORDER BY pr.id ASC
            """
        )
        for item in cursor.fetchall() or []:
            row_id_to_sheet_name[item["pattern_row_id"]] = item["sheet_name"]

        for image_row in image_rows:
            row_obj = row_obj_by_id.get(image_row["pattern_row_id"])
            if not row_obj:
                continue

            image_src = image_row.get("image_src")
            if image_src:
                row_obj["__images"].append(image_src)

            compare_png = image_row.get("compare_png")
            if compare_png is not None and not isinstance(compare_png, (bytes, bytearray)):
                # mysql connector đôi lúc trả memoryview
                try:
                    compare_png = bytes(compare_png)
                except Exception:
                    compare_png = None

            rebuilt_search_index.append(
                {
                    "sheetName": row_id_to_sheet_name.get(image_row["pattern_row_id"], ""),
                    "excelRow": row_obj.get("__excelRow"),
                    "rowRef": row_obj,
                    "matchedImage": image_src,
                    "hash": int(image_row["image_hash"]) if image_row.get("image_hash") not in (None, "") else 0,
                    "compare_png": compare_png,
                }
            )

        # 5) rebuild headers cho từng sheet
        known_header_order = [
            "No",
            "Customer",
            "Season",
            "Staff",
            "Style No",
            "Style Name",
            "Product",
            "Categories",
            "Gender",
            "Sketch Design",
        ]

        for import_id, sheets in sheets_by_import.items():
            for sheet in sheets:
                header_set = []
                seen = set()

                # thêm các cột chuẩn nếu row có dữ liệu
                for header in known_header_order:
                    real_header = "Sketch Design" if header == "Sketch Design" else header

                    has_value = False
                    if real_header == "Sketch Design":
                        has_value = any(row.get("__images") for row in sheet["rows"])
                    else:
                        has_value = any(
                            row.get(real_header) not in (None, "", [])
                            for row in sheet["rows"]
                        )

                    if has_value and real_header not in seen:
                        seen.add(real_header)
                        header_set.append(real_header)

                # thêm cột động từ extra_json
                for row in sheet["rows"]:
                    for key in row.keys():
                        if key.startswith("__"):
                            continue
                        if key not in seen:
                            seen.add(key)
                            header_set.append(key)

                if any(row.get("__images") for row in sheet["rows"]) and "Sketch Design" not in seen:
                    header_set.append("Sketch Design")

                sheet["headers"] = header_set
                sheet["rowCount"] = len(sheet["rows"])

        # 6) rebuild workbook list
        for import_item in import_rows:
            all_sheets = sheets_by_import.get(import_item["id"], [])

            # tính lại partition info như lúc parse file
            partition_info = infer_partition_from_sheets(all_sheets)

            workbook_search_index = [
                item
                for item in rebuilt_search_index
                if any(item["rowRef"] is row for sheet in all_sheets for row in sheet["rows"])
            ]

            workbook_data = {
                "id": import_item["id"],
                "fileName": import_item["file_name"],
                "sheetCount": import_item["sheet_count"] or len(all_sheets),
                "totalRows": import_item["total_rows"] or sum(len(sheet["rows"]) for sheet in all_sheets),
                "imageIndexCount": import_item["image_index_count"] or len(workbook_search_index),
                "mappedFolderCount": sum(
                    1
                    for sheet in all_sheets
                    for row in sheet["rows"]
                    if row.get("__folderId")
                ),
                "folderImportName": None,
                "customer": partition_info.get("customer"),
                "season": partition_info.get("season"),
                "customerValues": partition_info.get("customerValues", []),
                "seasonValues": partition_info.get("seasonValues", []),
                "isAmbiguous": partition_info.get("isAmbiguous", False),
                "sheets": all_sheets,
                "search_index": workbook_search_index,
            }

            STATE["imports"].append(workbook_data)

        rebuild_global_search_index()

    finally:
        cursor.close()
        conn.close()

def restore_folder_state_from_disk() -> dict[str, Any]:
    """
    Quét lại toàn bộ uploads/folder_imports, build lại STATE['folder_index'],
    rồi remap toàn bộ workbook theo Style No.

    Lưu ý:
    - folder_id sẽ được tạo mới mỗi lần restore (vì build_folder_index đang dùng uuid mới)
    - nên bắt buộc phải remap lại __folderId / __detailUrl cho từng row
    """
    # 1) reset folder state trong RAM
    STATE["folder_index"] = {}
    STATE["folder_root_dir"] = []

    # 2) xóa toàn bộ mapping folder cũ trên các row để tránh link stale
    for workbook in STATE.get("imports", []):
        for sheet in workbook.get("sheets", []):
            for row in sheet.get("rows", []):
                row["__folderId"] = None
                row["__folderName"] = None
                row["__detailUrl"] = None

        workbook["mappedFolderCount"] = 0
        workbook["folderImportName"] = None

    # 3) nếu chưa có thư mục folder_imports thì thôi
    if not FOLDER_IMPORT_DIR.exists():
        return {
            "rootCount": 0,
            "folderCount": 0,
            "mappedCount": 0,
        }

    # 4) lấy tất cả root folder import đã lưu trên disk
    saved_roots = [p for p in FOLDER_IMPORT_DIR.iterdir() if p.is_dir()]

    # sắp theo thời gian tạo/sửa gần giống thứ tự import
    saved_roots.sort(key=lambda p: (p.stat().st_mtime, natural_sort_key(p.name)))

    merged_folder_index: dict[str, dict[str, Any]] = {}
    restored_roots: list[Path] = []

    # 5) build lại folder index cho từng batch import
    for saved_root in saved_roots:
        try:
            folder_index = build_folder_index(saved_root)
        except Exception:
            # bỏ qua batch lỗi, không làm crash toàn app
            continue

        if not folder_index:
            continue

        restored_roots.append(saved_root)
        merged_folder_index.update(folder_index)

    STATE["folder_root_dir"] = restored_roots
    STATE["folder_index"] = merged_folder_index

    # 6) remap lại toàn bộ workbook theo toàn bộ folder đã restore
    total_mapped = 0
    for workbook in STATE.get("imports", []):
        mapped_count = map_folders_to_rows_by_style_no(workbook, STATE["folder_index"])
        total_mapped += mapped_count

        # Vì code hiện tại không lưu rootFolderName vào DB,
        # chỉ có thể set gần đúng theo batch folder cuối cùng
        if restored_roots:
            workbook["folderImportName"] = restored_roots[-1].name

    return {
        "rootCount": len(restored_roots),
        "folderCount": len(STATE["folder_index"]),
        "mappedCount": total_mapped,
    }

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_header_key(header: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(header or "").strip().lower())

DB_FIELD_MAPPING: dict[str, str] = {
    "No": "row_no",
    "Customer": "customer",
    "Season": "season",
    "Staff": "staff",
    "Style No": "style_no",
    "Style Name": "style_name",
    "Product": "product",
    "Categories": "categories",
    "Gender": "gender",
}

EXCEL_HEADER_ALIASES: dict[str, str] = {
    # STT
    "no": "No",
    "stt": "No",
    "rowno": "No",
    "number": "No",

    # Customer
    "customer": "Customer",
    "customername": "Customer",
    "customercode": "Customer",

    # Season / Staff
    "season": "Season",
    "staff": "Staff",

    # Style
    "styleno": "Style No",
    "stylenumber": "Style No",
    "stylecode": "Style No",
    "stylename": "Style Name",

    # Product / Category
    "product": "Product",
    "category": "Categories",
    "categories": "Categories",

    # Gender
    "gender": "Gender",
    "gendermenwomen": "Gender",
    "genderwomenmen": "Gender",
    "gendermw": "Gender",

    # Sketch
    "sketchdesign": "Sketch Design",
    "sketch": "Sketch Design",
}

STANDARD_HEADERS: set[str] = set(DB_FIELD_MAPPING.keys()) | {"Sketch Design"}


def normalize_excel_header(header: str) -> str:
    raw = str(header or "").strip()
    key = normalize_header_key(raw)
    return EXCEL_HEADER_ALIASES.get(key, raw)


def merge_extra_data_into_row(row_data: dict[str, Any], extra_data: dict[str, Any]) -> None:
    """
    Gộp extra_json vào row_data, đồng thời normalize alias cũ về tên chuẩn.
    Tránh việc Gender Men/Women và Gender cùng xuất hiện.
    """
    for raw_key, value in (extra_data or {}).items():
        canonical_key = normalize_excel_header(raw_key)

        # Nếu extra_json chứa alias của cột chuẩn thì dồn về cột chuẩn
        if canonical_key in DB_FIELD_MAPPING:
            if row_data.get(canonical_key) in (None, "", []):
                row_data[canonical_key] = value
            continue

        # Sketch Design không render như cột text
        if canonical_key == "Sketch Design":
            continue

        row_data[canonical_key] = value


def row_has_meaningful_data(row_data: dict[str, Any], headers: list[str]) -> bool:
    """
    Chỉ coi row là hợp lệ nếu có dữ liệu ở cột nghiệp vụ.
    Không tính cột STT / No.
    """
    ignored_headers = {"no", "stt"}

    for header in headers:
        header_key = normalize_header_key(header)
        if header_key in ignored_headers:
            continue

        value = row_data.get(header)

        if isinstance(value, str):
            if value.strip() != "":
                return True
        elif value not in (None, "", []):
            return True

    return False

def detect_header_row(ws, scan_rows: int = 20, scan_cols: int = 50) -> int:
    best_row = 1
    best_count = -1
    max_row = min(ws.max_row, scan_rows)
    max_col = min(ws.max_column, scan_cols)

    for row_idx in range(1, max_row + 1):
        count = 0
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                count += 1
        if count > best_count:
            best_count = count
            best_row = row_idx

    return best_row

def natural_sort_key(value: str) -> list[Any]:
    parts = re.split(r"(\d+)", str(value))
    result: list[Any] = []
    for part in parts:
        if part.isdigit():
            result.append(int(part))
        else:
            result.append(part.lower())
    return result

def normalize_match_value(value: Any) -> str:
    return str(value or "").strip()

def rebuild_global_search_index() -> None:
    merged_index: list[dict[str, Any]] = []

    for workbook in STATE.get("imports", []):
        for item in workbook.get("search_index", []):
            merged_index.append(item)

    STATE["search_index"] = merged_index


def get_active_workbook() -> dict[str, Any] | None:
    imports = STATE.get("imports", [])
    if not imports:
        return None
    return imports[-1]


def get_public_imports_state() -> list[dict[str, Any]]:
    result = []

    for workbook in STATE.get("imports", []):
        public_item = {k: v for k, v in workbook.items() if k != "search_index"}
        result.append(public_item)

    return result


def extract_folder_match_keys(folder_name: str) -> list[str]:
    raw_name = normalize_match_value(folder_name)
    keys: list[str] = []

    def add_key(value: str) -> None:
        clean = normalize_match_value(value)
        if clean and clean not in keys:
            keys.append(clean)

    # 1) lấy nguyên tên folder
    add_key(raw_name)

    # 2) lấy phần đứng trước ngoặc
    main_part = re.split(r"\s*\(", raw_name, maxsplit=1)[0].strip()
    add_key(main_part)

    # 3) lấy tất cả phần nằm trong ngoặc
    bracket_parts = re.findall(r"\(([^()]+)\)", raw_name)
    for part in bracket_parts:
        add_key(part)

    return keys


def pil_to_data_url(image: Image.Image, max_width: int = 240) -> str:
    img = image.copy()

    if img.width > max_width:
        ratio = max_width / float(img.width)
        new_size = (max_width, max(1, int(img.height * ratio)))
        img = img.resize(new_size, Image.Resampling.LANCZOS)

    output = BytesIO()
    img.save(output, format="PNG", optimize=True)
    base64_string = base64.b64encode(output.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{base64_string}"


def normalize_image(image: Image.Image) -> Image.Image:
    image = ImageOps.exif_transpose(image)

    if image.mode in ("RGBA", "LA"):
        background = Image.new("RGBA", image.size, (255, 255, 255, 255))
        image = Image.alpha_composite(background, image.convert("RGBA")).convert("RGB")
    else:
        image = image.convert("RGB")

    bg = Image.new("RGB", image.size, (255, 255, 255))
    diff = ImageChops.difference(image, bg)
    bbox = diff.getbbox()

    if bbox:
        image = image.crop(bbox)

    return image


def load_normalized_image_from_bytes(image_bytes: bytes) -> Image.Image:
    with Image.open(BytesIO(image_bytes)) as img:
        normalized = normalize_image(img)
        return normalized.copy()


def build_compare_gray_png(image: Image.Image, size: tuple[int, int] = (128, 128)) -> bytes:
    compare_img = image.convert("L").resize(size, Image.Resampling.LANCZOS)
    output = BytesIO()
    compare_img.save(output, format="PNG", optimize=True)
    return output.getvalue()


def compute_dhash(image: Image.Image, hash_size: int = 16) -> int:
    gray = image.convert("L").resize((hash_size + 1, hash_size), Image.Resampling.LANCZOS)
    pixels = list(gray.getdata())

    value = 0
    for row in range(hash_size):
        row_offset = row * (hash_size + 1)
        for col in range(hash_size):
            left_px = pixels[row_offset + col]
            right_px = pixels[row_offset + col + 1]
            value = (value << 1) | (1 if left_px > right_px else 0)

    return value


def hamming_distance(left: int, right: int) -> int:
    return (left ^ right).bit_count()


def pixel_similarity(compare_a_png: bytes, compare_b_png: bytes) -> float:
    with Image.open(BytesIO(compare_a_png)) as img_a, Image.open(BytesIO(compare_b_png)) as img_b:
        diff = ImageChops.difference(img_a, img_b)
        mean = ImageStat.Stat(diff).mean[0] / 255.0
        similarity = 1.0 - mean
        return max(0.0, min(1.0, similarity))


def build_image_signature(image_bytes: bytes) -> dict[str, Any]:
    normalized = load_normalized_image_from_bytes(image_bytes)
    return {
        "hash": compute_dhash(normalized),
        "compare_png": build_compare_gray_png(normalized),
    }


def image_bytes_to_record(image_bytes: bytes) -> dict[str, Any]:
    normalized = load_normalized_image_from_bytes(image_bytes)
    return {
        "src": pil_to_data_url(normalized),
        "hash": compute_dhash(normalized),
        "compare_png": build_compare_gray_png(normalized),
    }


def extract_images_by_row(ws, import_id: str, sheet_name: str) -> dict[int, list[dict[str, Any]]]:
    images_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)

    for image_index, image in enumerate(getattr(ws, "_images", [])):
        try:
            row_index = image.anchor._from.row + 1
            image_bytes = image._data()

            record = save_excel_image_file(
                image_bytes=image_bytes,
                import_id=import_id,
                sheet_name=sheet_name,
                row_idx=row_index,
                image_index=len(images_by_row[row_index]),
            )
            images_by_row[row_index].append(record)
        except Exception:
            continue

    return images_by_row


def clean_row_for_search(row_data: dict[str, Any]) -> dict[str, Any]:
    clean = {}
    for key, value in row_data.items():
        if key.startswith("__"):
            continue
        clean[key] = value
    return clean

def clear_folder_state() -> None:
    folder_roots = STATE.get("folder_root_dir") or []

    if not isinstance(folder_roots, list):
        folder_roots = [folder_roots]

    for folder_root_dir in folder_roots:
        if folder_root_dir and Path(folder_root_dir).exists():
            shutil.rmtree(folder_root_dir, ignore_errors=True)

    STATE["folder_root_dir"] = []
    STATE["folder_index"] = {}

    for workbook in STATE.get("imports", []):
        for sheet in workbook.get("sheets", []):
            for row in sheet.get("rows", []):
                row["__folderId"] = None
                row["__folderName"] = None
                row["__detailUrl"] = None
        workbook["mappedFolderCount"] = 0
        workbook["folderImportName"] = None

def clear_all_state() -> None:
    clear_folder_state()
    STATE["imports"] = []
    STATE["search_index"] = []

def get_reset_options() -> dict[str, Any]:
    customer_map: dict[str, set[str]] = {}

    for workbook in STATE.get("imports", []):
        customer = str(workbook.get("customer") or "").strip() or "Chưa xác định"

        if customer not in customer_map:
            customer_map[customer] = set()

        season_values = workbook.get("seasonValues") or []
        if season_values:
            for season in season_values:
                season_text = str(season or "").strip() or "Chưa xác định"
                customer_map[customer].add(season_text)
        else:
            season = str(workbook.get("season") or "").strip() or "Chưa xác định"
            customer_map[customer].add(season)

    customers = sorted(customer_map.keys(), key=natural_sort_key)

    return {
        "customers": customers,
        "seasonsByCustomer": {
            customer: sorted(list(seasons), key=natural_sort_key)
            for customer, seasons in customer_map.items()
        },
    }


def get_public_workbook_state() -> dict[str, Any] | None:
    workbook = get_active_workbook()
    if not workbook:
        return None
    return {k: v for k, v in workbook.items() if k != "search_index"}

def parse_excel_file(file_path: Path) -> dict[str, Any]:
    # clear_folder_state()
    import_id = uuid.uuid4().hex

    workbook = load_workbook(file_path, data_only=True)
    all_sheets: list[dict[str, Any]] = []
    total_rows = 0
    search_index: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        header_row = detect_header_row(ws)
        headers: list[str] = []
        column_mappings: list[tuple[str, int]] = []
        seen_headers: set[str] = set()

        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(header_row, col_idx).value
            if header_value in (None, ""):
                continue

            normalized_header = normalize_excel_header(str(header_value).strip())
            column_mappings.append((normalized_header, col_idx))

            if normalized_header not in seen_headers:
                seen_headers.add(normalized_header)
                headers.append(normalized_header)

        if not headers:
            continue

        images_by_row = extract_images_by_row(ws, import_id=import_id, sheet_name=sheet_name)
        rows: list[dict[str, Any]] = []

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data: dict[str, Any] = {}

            for header, col_idx in column_mappings:
                cell_value = normalize_cell_value(ws.cell(row_idx, col_idx).value)

                # Nếu nhiều cột Excel cùng map về 1 tên chuẩn,
                # ưu tiên giữ giá trị có dữ liệu
                if header not in row_data:
                    row_data[header] = cell_value
                else:
                    old_value = row_data.get(header)
                    if old_value in (None, "", []) and cell_value not in (None, "", []):
                        row_data[header] = cell_value

            row_images = images_by_row.get(row_idx, [])
            row_data["__images"] = [item["src"] for item in row_images]
            row_data["__excelRow"] = row_idx
            row_data["__folderId"] = None
            row_data["__folderName"] = None
            row_data["__detailUrl"] = None

            has_meaningful_data = row_has_meaningful_data(row_data, headers)

            if row_images:
                has_meaningful_data = True

            if has_meaningful_data:
                rows.append(row_data)

                for image_item in row_images:
                    search_index.append(
                        {
                            "sheetName": sheet_name,
                            "excelRow": row_idx,
                            "rowRef": row_data,
                            "matchedImage": image_item["src"],
                            "hash": image_item["hash"],
                            "compare_png": image_item["compare_png"],
                        }
                    )

        total_rows += len(rows)
        all_sheets.append(
            {
                "sheetName": sheet_name,
                "headerRow": header_row,
                "headers": headers,
                "rowCount": len(rows),
                "rows": rows,
            }
        )

    partition_info = infer_partition_from_sheets(all_sheets)
    workbook_data = {
        "id": import_id,
        "fileName": file_path.name,
        "sheetCount": len(all_sheets),
        "totalRows": total_rows,
        "imageIndexCount": len(search_index),
        "mappedFolderCount": 0,
        "folderImportName": None,
        "customer": partition_info.get("customer"),
        "season": partition_info.get("season"),
        "customerValues": partition_info.get("customerValues", []),
        "seasonValues": partition_info.get("seasonValues", []),
        "isAmbiguous": partition_info.get("isAmbiguous", False),
        "sheets": all_sheets,
        "search_index": search_index,
    }

    STATE["imports"].append(workbook_data)
    rebuild_global_search_index()

    return {k: v for k, v in workbook_data.items() if k != "search_index"}


def compare_signature(query_signature: dict[str, Any], target_item: dict[str, Any]) -> float:
    hash_size = 16 * 16
    hash_similarity = 1.0 - (
        hamming_distance(query_signature["hash"], target_item["hash"]) / hash_size
    )
    img_similarity = pixel_similarity(query_signature["compare_png"], target_item["compare_png"])
    score = (hash_similarity * 0.6) + (img_similarity * 0.4)
    return max(0.0, min(1.0, score))


def normalize_relative_parts(relative_path: str) -> list[str]:
    parts = []
    for part in Path(relative_path).parts:
        clean = str(part).strip()
        if clean in ("", ".", ".."):
            continue
        parts.append(clean)
    return parts


def save_uploaded_folder_files(files, relative_paths: list[str], root_folder_name: str) -> Path:
    import_id = uuid.uuid4().hex
    save_root = FOLDER_IMPORT_DIR / import_id
    save_root.mkdir(parents=True, exist_ok=True)

    for uploaded_file, rel_path in zip(files, relative_paths):
        parts = normalize_relative_parts(rel_path)

        if root_folder_name and parts and parts[0] == root_folder_name:
            parts = parts[1:]

        if len(parts) < 2:
            continue

        target_path = save_root.joinpath(*parts)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        uploaded_file.save(target_path)

    return save_root

def build_folder_index(saved_root: Path) -> dict[str, dict[str, Any]]:
    folder_index: dict[str, dict[str, Any]] = {}

    folder_paths = [
        p for p in saved_root.rglob("*")
        if p.is_dir()
    ]

    for folder_path in sorted(
        folder_paths,
        key=lambda x: natural_sort_key(x.relative_to(saved_root).as_posix())
    ):
        all_files = [
            p for p in folder_path.rglob("*")
            if p.is_file()
        ]

        if not all_files:
            continue

        file_items = []
        for file_path in sorted(all_files, key=lambda x: natural_sort_key(x.relative_to(folder_path).as_posix())):
            ext = file_path.suffix.lower()

            file_items.append(
                {
                    "name": file_path.name,
                    "relPath": file_path.relative_to(folder_path).as_posix(),
                    "ext": ext,
                    "sizeKb": round(file_path.stat().st_size / 1024, 1),
                    "isImage": ext in IMAGE_EXTENSIONS,
                    "isPdf": ext in PDF_EXTENSIONS,
                }
            )

        folder_id = uuid.uuid4().hex
        folder_index[folder_id] = {
            "id": folder_id,
            "name": folder_path.relative_to(saved_root).as_posix(),
            "matchName": folder_path.name,
            "path": folder_path,
            "fileCount": len(file_items),
            "files": file_items,
        }

    return folder_index


def map_folders_to_rows_by_style_no(workbook_data: dict[str, Any], folder_index: dict[str, dict[str, Any]]) -> int:
    folder_by_key: dict[str, dict[str, Any]] = {}

    for folder in folder_index.values():
        match_keys = extract_folder_match_keys(folder.get("matchName") or folder["name"])

        for key in match_keys:
            # nếu trùng key thì folder import sau sẽ ghi đè folder cũ
            folder_by_key[key] = folder

    mapped_count = 0

    for sheet in workbook_data.get("sheets", []):
        for row in sheet.get("rows", []):
            style_no = normalize_match_value(row.get("Style No"))
            if not style_no:
                continue

            folder = folder_by_key.get(style_no)
            if not folder:
                continue

            row["__folderId"] = folder["id"]
            row["__folderName"] = folder["name"]
            row["__detailUrl"] = f"/folder/{folder['id']}"
            mapped_count += 1

    workbook_data["mappedFolderCount"] = mapped_count
    return mapped_count

def safe_resolve_file(base_dir: Path, rel_path: str) -> Path:
    target = (base_dir / rel_path).resolve()
    base_resolved = base_dir.resolve()

    if base_resolved not in target.parents and target != base_resolved:
        raise FileNotFoundError("Đường dẫn không hợp lệ")

    if not target.exists() or not target.is_file():
        raise FileNotFoundError("Không tìm thấy file")

    return target


def write_audit_log(
    action_type: str,
    action_label: str,
    target_type: str | None = None,
    target_value: str | None = None,
    details: dict | None = None,
) -> None:
    user = get_current_user()
    user_id = user["id"] if user else None
    user_email = user["email"] if user else None
    ip_address = request.headers.get("X-Forwarded-For", request.remote_addr)

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO audit_logs (
                user_id, user_email,
                action_type, action_label,
                target_type, target_value,
                details_json, ip_address
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                user_id,
                user_email,
                action_type,
                action_label,
                target_type,
                target_value,
                json.dumps(details, ensure_ascii=False) if details else None,
                ip_address,
            ),
        )
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def hard_reset_all_data() -> None:
    """
    Reset toàn bộ dữ liệu dùng chung của hệ thống:
    - Xóa DB dữ liệu import
    - Xóa folder upload trên disk
    - Xóa state trong RAM
    Không đụng vào bảng users và audit_logs.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

        # Xóa theo thứ tự từ bảng con -> bảng cha
        cursor.execute("DELETE FROM pattern_row_images")
        cursor.execute("DELETE FROM pattern_rows")
        cursor.execute("DELETE FROM excel_sheets")
        cursor.execute("DELETE FROM excel_imports")

        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        conn.commit()
    except Exception:
        conn.rollback()
        try:
            cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        except Exception:
            pass
        raise
    finally:
        cursor.close()
        conn.close()

    # Xóa state runtime
    clear_all_state()

    # Xóa dữ liệu folder / ảnh đã import trên disk
    for path_obj in [FOLDER_IMPORT_DIR, EXCEL_IMAGE_DIR]:
        try:
            if path_obj.exists():
                shutil.rmtree(path_obj, ignore_errors=True)
            path_obj.mkdir(parents=True, exist_ok=True)
        except Exception:
            # Không làm crash toàn reset nếu lỗi disk nhẹ
            pass


def _reload_shared_state_after_partial_reset() -> None:
    """
    Rebuild lại state dùng chung từ DB + folder trên disk.
    Không xóa folder import trên disk.
    """
    STATE["imports"] = []
    STATE["search_index"] = []
    STATE["folder_index"] = {}
    STATE["folder_root_dir"] = []

    restore_state_from_mysql()
    restore_folder_state_from_disk()


def _delete_imports_from_db(import_ids: list[str]) -> int:
    """
    Xóa thật các import theo import_id khỏi DB + ảnh Excel trên disk.
    Trả về số import đã xóa.
    """
    import_ids = [str(x).strip() for x in import_ids if str(x).strip()]
    if not import_ids:
        return 0

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        placeholders = ", ".join(["%s"] * len(import_ids))

        # Lấy pattern_row ids để xóa ảnh con trước
        cursor.execute(
            f"""
            SELECT id
            FROM pattern_rows
            WHERE import_id IN ({placeholders})
            """,
            tuple(import_ids),
        )
        pattern_row_ids = [row["id"] for row in (cursor.fetchall() or [])]

        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

        if pattern_row_ids:
            row_placeholders = ", ".join(["%s"] * len(pattern_row_ids))
            cursor.execute(
                f"""
                DELETE FROM pattern_row_images
                WHERE pattern_row_id IN ({row_placeholders})
                """,
                tuple(pattern_row_ids),
            )

        cursor.execute(
            f"""
            DELETE FROM pattern_rows
            WHERE import_id IN ({placeholders})
            """,
            tuple(import_ids),
        )

        cursor.execute(
            f"""
            DELETE FROM excel_sheets
            WHERE import_id IN ({placeholders})
            """,
            tuple(import_ids),
        )

        cursor.execute(
            f"""
            DELETE FROM excel_imports
            WHERE id IN ({placeholders})
            """,
            tuple(import_ids),
        )

        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        conn.commit()

    except Exception:
        conn.rollback()
        try:
            cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        except Exception:
            pass
        raise
    finally:
        cursor.close()
        conn.close()

    # Xóa ảnh Excel trên disk theo từng import
    for import_id in import_ids:
        try:
            image_dir = EXCEL_IMAGE_DIR / import_id
            if image_dir.exists():
                shutil.rmtree(image_dir, ignore_errors=True)
        except Exception:
            pass

    # Rebuild lại state dùng chung từ DB + folder disk
    _reload_shared_state_after_partial_reset()

    return len(import_ids)


def _hard_reset_rows(
    row_where_sql: str,
    row_where_params: tuple[Any, ...],
    joined_where_sql: str,
    joined_where_params: tuple[Any, ...],
) -> dict[str, int]:
    """
    Xóa dữ liệu theo từng row trong DB, không xóa theo cấp file/import.
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS row_count
            FROM pattern_rows
            WHERE {row_where_sql}
            """,
            row_where_params,
        )
        removed_row_count = int((cursor.fetchone() or {}).get("row_count") or 0)

        if removed_row_count == 0:
            return {
                "removedRowCount": 0,
                "affectedImportCount": 0,
                "removedImportCount": 0,
            }

        cursor.execute(
            f"""
            SELECT DISTINCT import_id
            FROM pattern_rows
            WHERE {row_where_sql}
            """,
            row_where_params,
        )
        affected_import_ids = [
            str(row["import_id"])
            for row in (cursor.fetchall() or [])
            if row.get("import_id")
        ]

        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

        # Xóa ảnh của các row sẽ bị xóa
        cursor.execute(
            f"""
            DELETE pri
            FROM pattern_row_images pri
            INNER JOIN pattern_rows pr ON pr.id = pri.pattern_row_id
            WHERE {joined_where_sql}
            """,
            joined_where_params,
        )

        # Xóa row đúng customer / season
        cursor.execute(
            f"""
            DELETE FROM pattern_rows
            WHERE {row_where_sql}
            """,
            row_where_params,
        )

        removed_import_ids: list[str] = []

        if affected_import_ids:
            placeholders = ", ".join(["%s"] * len(affected_import_ids))

            # Xóa sheet rỗng sau khi row đã bị xóa
            cursor.execute(
                f"""
                DELETE es
                FROM excel_sheets es
                LEFT JOIN pattern_rows pr ON pr.sheet_id = es.id
                WHERE es.import_id IN ({placeholders})
                  AND pr.id IS NULL
                """,
                tuple(affected_import_ids),
            )

            # Cập nhật lại metadata từng import
            for import_id in affected_import_ids:
                cursor.execute(
                    """
                    SELECT COUNT(*) AS total_rows
                    FROM pattern_rows
                    WHERE import_id = %s
                    """,
                    (import_id,),
                )
                total_rows = int((cursor.fetchone() or {}).get("total_rows") or 0)

                cursor.execute(
                    """
                    SELECT COUNT(*) AS sheet_count
                    FROM excel_sheets
                    WHERE import_id = %s
                    """,
                    (import_id,),
                )
                sheet_count = int((cursor.fetchone() or {}).get("sheet_count") or 0)

                cursor.execute(
                    """
                    SELECT COUNT(*) AS image_index_count
                    FROM pattern_row_images pri
                    INNER JOIN pattern_rows pr ON pr.id = pri.pattern_row_id
                    WHERE pr.import_id = %s
                    """,
                    (import_id,),
                )
                image_index_count = int((cursor.fetchone() or {}).get("image_index_count") or 0)

                if total_rows <= 0:
                    removed_import_ids.append(import_id)
                else:
                    cursor.execute(
                        """
                        UPDATE excel_imports
                        SET
                            total_rows = %s,
                            sheet_count = %s,
                            image_index_count = %s
                        WHERE id = %s
                        """,
                        (total_rows, sheet_count, image_index_count, import_id),
                    )

            if removed_import_ids:
                placeholders_removed = ", ".join(["%s"] * len(removed_import_ids))
                cursor.execute(
                    f"""
                    DELETE FROM excel_imports
                    WHERE id IN ({placeholders_removed})
                    """,
                    tuple(removed_import_ids),
                )

        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        conn.commit()

    except Exception:
        conn.rollback()
        try:
            cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        except Exception:
            pass
        raise
    finally:
        cursor.close()
        conn.close()

    # Nếu import nào đã trống hoàn toàn thì xóa luôn folder ảnh của import đó
    for import_id in removed_import_ids:
        try:
            image_dir = EXCEL_IMAGE_DIR / import_id
            if image_dir.exists():
                shutil.rmtree(image_dir, ignore_errors=True)
        except Exception:
            pass

    _reload_shared_state_after_partial_reset()

    return {
        "removedRowCount": removed_row_count,
        "affectedImportCount": len(affected_import_ids),
        "removedImportCount": len(removed_import_ids),
    }


def hard_reset_by_customer(customer: str) -> dict[str, int]:
    customer = str(customer or "").strip()
    if not customer:
        return {
            "removedRowCount": 0,
            "affectedImportCount": 0,
            "removedImportCount": 0,
        }

    return _hard_reset_rows(
        row_where_sql="customer = %s",
        row_where_params=(customer,),
        joined_where_sql="pr.customer = %s",
        joined_where_params=(customer,),
    )


def hard_reset_by_customer_and_season(customer: str, season: str) -> dict[str, int]:
    customer = str(customer or "").strip()
    season = str(season or "").strip()

    if not customer or not season:
        return {
            "removedRowCount": 0,
            "affectedImportCount": 0,
            "removedImportCount": 0,
        }

    return _hard_reset_rows(
        row_where_sql="customer = %s AND season = %s",
        row_where_params=(customer, season),
        joined_where_sql="pr.customer = %s AND pr.season = %s",
        joined_where_params=(customer, season),
    )



@app.get("/login")
def login_page():
    return render_template("login.html")


@app.post("/api/login")
def api_login():
    payload = request.get_json(silent=True) or request.form

    email = normalize_email(payload.get("email"))
    password = str(payload.get("password", ""))

    if not email or not password:
        return jsonify({"success": False, "message": "Thiếu email hoặc mật khẩu"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT * FROM users WHERE email = %s AND is_active = 1",
            (email,),
        )
        user = cursor.fetchone()

        if not user or not check_password_hash(user["password_hash"], password):
            return jsonify({"success": False, "message": "Sai email hoặc mật khẩu"}), 401

        session["user_id"] = user["id"]
        session["user_email"] = user["email"]

        return jsonify(
            {
                "success": True,
                "message": "Đăng nhập thành công",
                "data": user_to_public_payload(user),
            }
        )
    finally:
        cursor.close()
        conn.close()


@app.post("/api/logout")
def api_logout():
    session.clear()
    return jsonify({"success": True, "message": "Đã đăng xuất"})

@app.get("/")
@login_required_page
def index():
    return render_template("index.html", base_path="")

@app.get("/api/me")
@login_required_api
def api_me():
    user = get_current_user()
    return jsonify({"success": True, "data": user_to_public_payload(user)})


@app.get("/admin")
@login_required_page
@permission_required_page("can_manage_users")
def admin_page():
    return render_template("admin.html")

@app.get("/api/admin/users")
@login_required_api
@permission_required_api("can_manage_users")
def admin_list_users():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT
            id, email, full_name, is_active, is_admin,
            can_import_excel, can_import_folder, can_search_image,
            can_view_data, can_reset_data, can_manage_users,
            can_view_audit_logs,
            created_at
            FROM users
            ORDER BY created_at DESC, id DESC
            """
        )
        users = cursor.fetchall() or []

        return jsonify(
            {
                "success": True,
                "data": {
                    "users": [
                        {
                            "id": user["id"],
                            "email": user["email"],
                            "fullName": user.get("full_name") or "",
                            "isActive": bool(user["is_active"]),
                            "isAdmin": bool(user["is_admin"]),
                            "permissions": {
                                "canImportExcel": bool(user["can_import_excel"]),
                                "canImportFolder": bool(user["can_import_folder"]),
                                "canSearchImage": bool(user["can_search_image"]),
                                "canViewData": bool(user["can_view_data"]),
                                "canResetData": bool(user["can_reset_data"]),
                                "canManageUsers": bool(user["can_manage_users"]),
                                "canViewAuditLogs": bool(user["can_view_audit_logs"]),
                            },
                            "createdAt": str(user["created_at"]),
                        }
                        for user in users
                    ]
                },
            }
        )
    finally:
        cursor.close()
        conn.close()


@app.post("/api/admin/users")
@login_required_api
@permission_required_api("can_manage_users")
def admin_create_user():
    payload = request.get_json(silent=True) or request.form

    email = normalize_email(payload.get("email"))
    password = str(payload.get("password", "")).strip()
    full_name = str(payload.get("full_name", "")).strip()

    if not email or not password:
        return jsonify({"success": False, "message": "Thiếu email hoặc mật khẩu"}), 400

    password_hash = generate_password_hash(password)

    can_import_excel = int(bool(payload.get("can_import_excel", True)))
    can_import_folder = int(bool(payload.get("can_import_folder", True)))
    can_search_image = int(bool(payload.get("can_search_image", True)))
    can_view_data = int(bool(payload.get("can_view_data", True)))
    can_reset_data = int(bool(payload.get("can_reset_data", False)))
    can_manage_users = int(bool(payload.get("can_manage_users", False)))
    is_admin = int(bool(payload.get("is_admin", False)))
    is_active = int(bool(payload.get("is_active", True)))
    can_view_audit_logs = int(bool(payload.get("can_view_audit_logs", False)))

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO users (
            email, password_hash, full_name, is_active, is_admin,
            can_import_excel, can_import_folder, can_search_image,
            can_view_data, can_reset_data, can_manage_users,
            can_view_audit_logs
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                email,
                password_hash,
                full_name,
                is_active,
                is_admin,
                can_import_excel,
                can_import_folder,
                can_search_image,
                can_view_data,
                can_reset_data,
                can_manage_users,
                can_view_audit_logs,
            ),
        )
        conn.commit()

        return jsonify({"success": True, "message": "Tạo tài khoản thành công"})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "message": f"Không tạo được tài khoản: {error}"}), 500
    finally:
        cursor.close()
        conn.close()

@app.put("/api/admin/users/<int:user_id>/permissions")
@login_required_api
@permission_required_api("can_manage_users")
def admin_update_user_permissions(user_id: int):
    payload = request.get_json(silent=True) or {}

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            UPDATE users
            SET
            full_name = %s,
            is_active = %s,
            is_admin = %s,
            can_import_excel = %s,
            can_import_folder = %s,
            can_search_image = %s,
            can_view_data = %s,
            can_reset_data = %s,
            can_manage_users = %s,
            can_view_audit_logs = %s
            WHERE id = %s
            """,
            (
                str(payload.get("full_name", "")).strip(),
                int(bool(payload.get("is_active", True))),
                int(bool(payload.get("is_admin", False))),
                int(bool(payload.get("can_import_excel", True))),
                int(bool(payload.get("can_import_folder", True))),
                int(bool(payload.get("can_search_image", True))),
                int(bool(payload.get("can_view_data", True))),
                int(bool(payload.get("can_reset_data", False))),
                int(bool(payload.get("can_manage_users", False))),
                int(bool(payload.get("can_view_audit_logs", False))),
                user_id,
            ),
        )
        conn.commit()

        return jsonify({"success": True, "message": "Cập nhật quyền thành công"})
    finally:
        cursor.close()
        conn.close()

@app.put("/api/admin/users/<int:user_id>/password")
@login_required_api
@permission_required_api("can_manage_users")
def admin_reset_user_password(user_id: int):
    payload = request.get_json(silent=True) or {}
    new_password = str(payload.get("password", "")).strip()

    if not new_password:
        return jsonify({"success": False, "message": "Thiếu mật khẩu mới"}), 400

    password_hash = generate_password_hash(new_password)

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE users SET password_hash = %s WHERE id = %s",
            (password_hash, user_id),
        )
        conn.commit()

        return jsonify({"success": True, "message": "Đổi mật khẩu thành công"})
    finally:
        cursor.close()
        conn.close()


@app.get("/media/<path:rel_path>")
def media_file(rel_path: str):
    try:
        file_path = safe_resolve_file(UPLOAD_DIR, rel_path)
        return send_file(file_path, as_attachment=False, download_name=file_path.name)
    except Exception:
        abort(404)

@app.get("/folder/<folder_id>")
def folder_detail(folder_id: str):
    folder = STATE["folder_index"].get(folder_id)
    if not folder:
        abort(404)

    return render_template("folder_detail.html", folder=folder)

@app.get("/folder-file/<folder_id>/<path:rel_path>")
def folder_file(folder_id: str, rel_path: str):
    folder = STATE["folder_index"].get(folder_id)
    if not folder:
        abort(404)

    try:
        file_path = safe_resolve_file(folder["path"], rel_path)
        return send_file(file_path, as_attachment=False, download_name=file_path.name)
    except Exception:
        abort(404)

@app.get(f"/folder-download/<folder_id>")
def folder_download(folder_id: str):
    folder = STATE["folder_index"].get(folder_id)
    if not folder:
        abort(404)

    memory_file = BytesIO()

    with zipfile.ZipFile(memory_file, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(folder["path"].rglob("*"), key=lambda x: natural_sort_key(x.as_posix())):
            if not file_path.is_file():
                continue

            arcname = f"{folder['name']}/{file_path.relative_to(folder['path']).as_posix()}"
            zf.write(file_path, arcname)

    memory_file.seek(0)

    zip_name = secure_filename(folder["name"]) or "folder"
    return send_file(
        memory_file,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{zip_name}.zip",
    )

@app.get(f"/folder-download-file/<folder_id>/<path:rel_path>")
def folder_download_file(folder_id: str, rel_path: str):
    folder = STATE["folder_index"].get(folder_id)
    if not folder:
        abort(404)

    try:
        file_path = safe_resolve_file(folder["path"], rel_path)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=file_path.name,
        )
    except Exception:
        abort(404)

@app.post("/api/excel/upload")
def upload_excel():
    uploaded_file = request.files.get("file")

    if not uploaded_file or uploaded_file.filename == "":
        return jsonify({"success": False, "message": "Bạn chưa chọn file Excel"}), 400

    if not allowed_file(uploaded_file.filename):
        return jsonify({"success": False, "message": "Chỉ hỗ trợ file .xlsx hoặc .xlsm"}), 400

    safe_name = secure_filename(uploaded_file.filename)
    file_path = UPLOAD_DIR / safe_name
    uploaded_file.save(file_path)

    try:
        result = parse_excel_file(file_path)

        full_workbook = next(
            (item for item in STATE["imports"] if item["id"] == result["id"]),
            None
        )
        if full_workbook:
            save_workbook_to_mysql(full_workbook)

        write_audit_log(
            action_type="IMPORT_EXCEL",
            action_label="Import Excel",
            target_type="excel_file",
            target_value=uploaded_file.filename,
            details={
                "workbook_id": result.get("id"),
                "file_name": result.get("fileName"),
                "sheet_count": result.get("sheetCount"),
                "total_rows": result.get("totalRows"),
                "image_index_count": result.get("imageIndexCount"),
            },
        )

        return jsonify({
            "success": True,
            "message": "Import Excel thành công, dữ liệu mới đã được cộng dồn",
            "data": {
                "workbook": result,
                "imports": get_public_imports_state(),
            }
        })
    except Exception as error:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Không đọc được file Excel: {error}"}), 500
    finally:
        if file_path.exists():
            file_path.unlink(missing_ok=True)

@app.post("/api/folder/import")
def import_local_folder():
    workbook = get_active_workbook()
    if not workbook:
        return jsonify({"success": False, "message": "Bạn cần import Excel trước"}), 400

    files = request.files.getlist("files")
    relative_paths = request.form.getlist("relativePaths")
    root_folder_name = str(request.form.get("rootFolderName", "")).strip()

    if not files:
        return jsonify({"success": False, "message": "Bạn chưa chọn folder"}), 400

    if len(files) != len(relative_paths):
        return jsonify({"success": False, "message": "Dữ liệu folder upload không khớp"}), 400

    try:
        # KHÔNG clear_folder_state() ở đây nữa
        saved_root = save_uploaded_folder_files(files, relative_paths, root_folder_name)

        all_dirs = sorted(
            [p.relative_to(saved_root).as_posix() for p in saved_root.rglob("*") if p.is_dir()],
            key=natural_sort_key
        )
        all_files = sorted(
            [p.relative_to(saved_root).as_posix() for p in saved_root.rglob("*") if p.is_file()],
            key=natural_sort_key
        )

        print("=== DEBUG FOLDER IMPORT ===")
        print("Uploaded file count:", len(files))
        print("Relative paths from browser:")
        for path in relative_paths:
            print(" -", path)

        print("Directories actually saved on server:", len(all_dirs))
        for d in all_dirs:
            print(" [DIR]", d)

        print("Files actually saved on server:", len(all_files))
        for f in all_files:
            print(" [FILE]", f)
        print("=== END DEBUG ===")

        new_folder_index = build_folder_index(saved_root)

        # đổi folder_root_dir thành list để giữ nhiều lần import
        existing_roots = STATE.get("folder_root_dir") or []
        if not isinstance(existing_roots, list):
            existing_roots = [existing_roots]
        existing_roots.append(saved_root)
        STATE["folder_root_dir"] = existing_roots

        # gộp folder mới vào folder cũ
        merged_folder_index = dict(STATE.get("folder_index", {}))
        merged_folder_index.update(new_folder_index)
        STATE["folder_index"] = merged_folder_index

        # map lại toàn bộ workbook theo toàn bộ folder đã có
        mapped_count = 0
        for workbook_item in STATE.get("imports", []):
            mapped_count += map_folders_to_rows_by_style_no(workbook_item, STATE["folder_index"])
            workbook_item["folderImportName"] = root_folder_name or saved_root.name

        write_audit_log(
            action_type="IMPORT_FOLDER",
            action_label="Import Folder",
            target_type="folder_root",
            target_value=root_folder_name or saved_root.name,
            details={
                "mapped_count": mapped_count,
                "folder_count": len(new_folder_index),
                "uploaded_file_count": len(files),
            },
        )


        return jsonify(
            {
                "success": True,
                "message": "Import folder thành công",
                "data": {
                    "folderCount": len(STATE["folder_index"]),      # tổng folder hiện có
                    "addedFolderCount": len(new_folder_index),      # số folder vừa thêm
                    "mappedCount": mapped_count,
                    "workbook": get_public_workbook_state(),
                    "imports": get_public_imports_state(),
                },
            }
        )
    except Exception as error:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Không import được folder: {error}"}), 500

@app.post("/api/search-by-image")
def search_by_image():
    if not STATE["search_index"]:
        return jsonify(
            {
                "success": False,
                "message": "Bạn cần import Excel trước khi dán ảnh để tìm kiếm",
            }
        ), 400

    pasted_image = request.files.get("image")
    if not pasted_image:
        return jsonify({"success": False, "message": "Không nhận được ảnh từ clipboard"}), 400

    try:
        image_bytes = pasted_image.read()
        query_signature = build_image_signature(image_bytes)

        results = []
        for item in STATE["search_index"]:
            score = compare_signature(query_signature, item)
            row_ref = item["rowRef"]

            results.append(
                {
                    "sheetName": item["sheetName"],
                    "excelRow": item["excelRow"],
                    "score": round(score, 4),
                    "scorePercent": int(round(score * 100)),
                    "matchedImage": item["matchedImage"],
                    "detailUrl": row_ref.get("__detailUrl"),
                    "folderName": row_ref.get("__folderName"),
                    "row": clean_row_for_search(row_ref),
                }
            )

        results.sort(key=lambda x: x["score"], reverse=True)

        return jsonify(
            {
                "success": True,
                "message": "So khớp ảnh thành công",
                "data": {
                    "results": results[:10],
                },
            }
        )
    except Exception as error:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Lỗi xử lý ảnh: {error}"}), 500


@app.get("/api/state")
def get_state():
    if not STATE["imports"]:
        restore_state_from_mysql()

    if STATE["imports"] and not STATE["folder_index"]:
        restore_folder_state_from_disk()

    imports = get_public_imports_state()
    active_workbook = get_public_workbook_state()

    return jsonify(
        {
            "success": True,
            "data": {
                "imports": imports,
                "workbook": active_workbook,
                "hasWorkbook": len(imports) > 0,
                "importCount": len(imports),
                "totalImageIndexCount": len(STATE.get("search_index", [])),
                "resetOptions": get_reset_options(),
            },
        }
    )

@app.post("/api/reset")
def reset_state():
    try:
        hard_reset_all_data()

        write_audit_log(
            action_type="RESET_ALL",
            action_label="Reset toàn bộ dữ liệu",
            target_type="system",
            target_value="all_data",
        )

        return jsonify(
            {
                "success": True,
                "message": "Đã reset toàn bộ dữ liệu",
            }
        )
    except Exception as error:
        import traceback
        traceback.print_exc()
        return jsonify(
            {"success": False, "message": f"Không reset được dữ liệu: {error}"}
        ), 500
    
@app.post("/api/reset/customer")
def reset_customer_data():
    payload = request.get_json(silent=True) or request.form
    customer = str(payload.get("customer", "")).strip()

    if not customer:
        return jsonify({"success": False, "message": "Bạn chưa chọn customer"}), 400

    reset_result = hard_reset_by_customer(customer)

    if reset_result["removedRowCount"] == 0:
        return jsonify({"success": False, "message": "Không có dữ liệu để reset"}), 404

    write_audit_log(
        action_type="RESET_CUSTOMER",
        action_label="Reset theo customer",
        target_type="customer",
        target_value=customer,
        details={
            "customer": customer,
            "removed_row_count": reset_result["removedRowCount"],
            "affected_import_count": reset_result["affectedImportCount"],
            "removed_import_count": reset_result["removedImportCount"],
        },
    )

    return jsonify(
        {
            "success": True,
            "message": (
                f"Đã reset {reset_result['removedRowCount']} dòng dữ liệu của customer {customer}"
            ),
            "data": {
                "imports": get_public_imports_state(),
                "resetOptions": get_reset_options(),
            },
        }
    )

@app.post("/api/reset/season")
def reset_customer_season_data():
    payload = request.get_json(silent=True) or request.form
    customer = str(payload.get("customer", "")).strip()
    season = str(payload.get("season", "")).strip()

    if not customer or not season:
        return jsonify({"success": False, "message": "Bạn cần chọn customer và season"}), 400

    reset_result = hard_reset_by_customer_and_season(customer, season)

    if reset_result["removedRowCount"] == 0:
        return jsonify({"success": False, "message": "Không có dữ liệu để reset"}), 404

    write_audit_log(
        action_type="RESET_SEASON",
        action_label="Reset theo customer / season",
        target_type="customer_season",
        target_value=f"{customer} / {season}",
        details={
            "customer": customer,
            "season": season,
            "removed_row_count": reset_result["removedRowCount"],
            "affected_import_count": reset_result["affectedImportCount"],
            "removed_import_count": reset_result["removedImportCount"],
        },
    )

    return jsonify(
        {
            "success": True,
            "message": (
                f"Đã reset {reset_result['removedRowCount']} dòng dữ liệu của {customer} / {season}"
            ),
            "data": {
                "imports": get_public_imports_state(),
                "resetOptions": get_reset_options(),
            },
        }
    )


@app.get("/api/admin/audit-logs")
@login_required_api
@permission_required_api("can_view_audit_logs")
def admin_audit_logs():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            """
            SELECT
                id,
                user_id,
                user_email,
                action_type,
                action_label,
                target_type,
                target_value,
                details_json,
                ip_address,
                created_at
            FROM audit_logs
            ORDER BY created_at DESC, id DESC
            LIMIT 300
            """
        )
        rows = cursor.fetchall() or []

        return jsonify(
            {
                "success": True,
                "data": {
                    "logs": rows
                },
            }
        )
    finally:
        cursor.close()
        conn.close()


# email = "admin@gmail.com"
# raw_password = "123456"
# password_hash = generate_password_hash(raw_password)

if __name__ == "__main__":
    restore_state_from_mysql()
    restore_folder_state_from_disk()
    app.run(debug=True)
 

    





  